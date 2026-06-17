from datetime import date
from io import BytesIO
from typing import Any

from fastapi import APIRouter, BackgroundTasks, Depends, Query
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.tenant import CurrentUser, get_current_user
from app.database import get_session

router = APIRouter(prefix="/reports", tags=["Relatórios"])

_HDR = "1A3F6F"


def _mk(sheet: str):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    return wb, ws


def _headers(ws, cols: list[str]):
    for i, h in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22


def _widths(ws):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(w + 3, 48)


def _xlsx(wb: Workbook, name: str) -> Response:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{name}"'},
    )


import re as _re
_ILLEGAL = _re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

def _s(v: Any) -> str:
    s = str(v) if v is not None else "—"
    return _ILLEGAL.sub('', s)


# ─── Financeiro ───────────────────────────────────────────────────────────────

async def _query_finance(session, aid: str, date_from=None, date_to=None, tx_type=None, payment_method=None):
    conds = ["t.association_id = :aid", "t.reversed_at IS NULL"]
    p: dict = {"aid": aid}
    if date_from: conds.append("t.created_at::date >= :df"); p["df"] = date.fromisoformat(date_from)
    if date_to: conds.append("t.created_at::date <= :dt"); p["dt"] = date.fromisoformat(date_to)
    if tx_type in ("income", "expense"): conds.append("t.type = :tp"); p["tp"] = tx_type
    if payment_method: conds.append("pm.name ILIKE :pm"); p["pm"] = f"%{payment_method}%"
    w = " AND ".join(conds)
    rows = (await session.execute(text(f"""
        SELECT t.created_at::date, t.type, t.amount, t.description,
               pm.name AS pagamento, u.full_name AS criado_por
        FROM transactions t
        LEFT JOIN payment_methods pm ON pm.id = t.payment_method_id
        LEFT JOIN users u ON u.id = t.created_by
        WHERE {w} ORDER BY t.created_at DESC
    """), p)).fetchall()
    return [{"Data": _s(r[0]), "Tipo": "Entrada" if r[1] == "income" else "Saída",
             "Valor (R$)": float(r[2] or 0), "Descrição": _s(r[3]),
             "Pagamento": _s(r[4]), "Criado por": _s(r[5])} for r in rows]


@router.get("/finance/preview")
async def preview_finance(
    date_from: str | None = None, date_to: str | None = None,
    tx_type: str | None = None, payment_method: str | None = None,
    current: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> list[dict]:
    return await _query_finance(session, str(current.association_id), date_from, date_to, tx_type, payment_method)


@router.get("/finance")
async def export_finance(
    date_from: str | None = None, date_to: str | None = None,
    tx_type: str | None = None, payment_method: str | None = None,
    current: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> Response:
    rows = await _query_finance(session, str(current.association_id), date_from, date_to, tx_type, payment_method)
    wb, ws = _mk("Financeiro")
    cols = list(rows[0].keys()) if rows else ["Data","Tipo","Valor (R$)","Descrição","Pagamento","Criado por"]
    _headers(ws, cols)
    for r in rows: ws.append(list(r.values()))
    _widths(ws)
    return _xlsx(wb, f"financeiro.xlsx")


# ─── Moradores ────────────────────────────────────────────────────────────────

async def _query_residents(session, aid: str, res_type=None, res_status=None, q=None):
    conds = ["association_id = :aid", "status != 'suspended'"]
    p: dict = {"aid": aid}
    if res_type: conds.append("type = :tp"); p["tp"] = res_type
    if res_status: conds.append("status = :st"); p["st"] = res_status
    if q: conds.append("(unaccent(lower(full_name)) LIKE unaccent(lower(:q)) OR cpf ILIKE :q)"); p["q"] = f"%{q}%"
    w = " AND ".join(conds)
    rows = (await session.execute(text(f"""
        SELECT full_name, cpf, phone_primary, email, address_street,
               address_number, address_neighborhood, address_cep,
               type, status, created_at::date
        FROM residents WHERE {w} ORDER BY full_name
    """), p)).fetchall()
    cols = ["Nome","CPF","Telefone","E-mail","Rua","Número","Bairro","CEP","Tipo","Status","Cadastrado em"]
    return [dict(zip(cols, [_s(v) for v in r])) for r in rows]


@router.get("/residents/preview")
async def preview_residents(
    res_type: str | None = None, res_status: str | None = None, q: str | None = None,
    current: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> list[dict]:
    return await _query_residents(session, str(current.association_id), res_type, res_status, q)


@router.get("/residents")
async def export_residents(
    res_type: str | None = None, res_status: str | None = None, q: str | None = None,
    current: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> Response:
    rows = await _query_residents(session, str(current.association_id), res_type, res_status, q)
    wb, ws = _mk("Moradores")
    cols = list(rows[0].keys()) if rows else []
    _headers(ws, cols)
    for r in rows: ws.append(list(r.values()))
    _widths(ws)
    return _xlsx(wb, "moradores.xlsx")


# ─── Encomendas ───────────────────────────────────────────────────────────────

async def _query_packages(session, aid: str, date_from=None, date_to=None, pkg_status=None, operator_ids=None, street=None, cep=None):
    conds = ["p.association_id = :aid"]
    p: dict = {"aid": aid}
    if pkg_status == 'awaiting':
        conds.append("p.status IN ('received', 'notified')")
        # ignora janela de data — mostra TODOS aguardando independente de quando chegaram
    else:
        if date_from: conds.append("p.received_at::date >= :df"); p["df"] = date.fromisoformat(date_from)
        if date_to: conds.append("p.received_at::date <= :dt"); p["dt"] = date.fromisoformat(date_to)
        if pkg_status: conds.append("p.status = :st"); p["st"] = pkg_status
    if operator_ids:
        placeholders = ", ".join(f":op{i}" for i in range(len(operator_ids)))
        conds.append(f"p.received_by::text IN ({placeholders})")
        for i, oid in enumerate(operator_ids): p[f"op{i}"] = oid
    if street:
        conds.append("r.address_street ILIKE :street")
        p["street"] = f"%{street}%"
    if cep:
        conds.append("regexp_replace(r.address_cep, '[^0-9]', '', 'g') = regexp_replace(:cep, '[^0-9]', '', 'g')")
        p["cep"] = cep
    w = " AND ".join(conds)
    rows = (await session.execute(text(f"""
        SELECT p.tracking_code, r.full_name, r.address_street, r.address_cep,
               p.status, p.carrier_name, p.received_at::date, p.delivered_at::date,
               p.delivery_fee_amount, u.full_name
        FROM packages p
        LEFT JOIN residents r ON r.id = p.resident_id
        LEFT JOIN users u ON u.id = p.received_by
        WHERE {w} ORDER BY p.received_at DESC
    """), p)).fetchall()
    cols = ["Código Rastreio","Destinatário","Rua","CEP","Status","Transportadora",
            "Recebido em","Entregue em","Taxa (R$)","Recebido por"]
    return [dict(zip(cols, [_s(v) for v in r])) for r in rows]


@router.get("/packages/preview")
async def preview_packages(
    date_from: str | None = None, date_to: str | None = None,
    pkg_status: str | None = None,
    operator_ids: list[str] | None = Query(default=None),
    street: str | None = None, cep: str | None = None,
    current: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> list[dict]:
    return await _query_packages(session, str(current.association_id), date_from, date_to, pkg_status, operator_ids, street, cep)


@router.get("/packages")
async def export_packages(
    date_from: str | None = None, date_to: str | None = None,
    pkg_status: str | None = None,
    operator_ids: list[str] | None = Query(default=None),
    street: str | None = None, cep: str | None = None,
    current: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> Response:
    rows = await _query_packages(session, str(current.association_id), date_from, date_to, pkg_status, operator_ids, street, cep)
    wb, ws = _mk("Encomendas")
    cols = list(rows[0].keys()) if rows else []
    _headers(ws, cols)
    for r in rows: ws.append(list(r.values()))
    _widths(ws)
    return _xlsx(wb, "encomendas.xlsx")


# ─── Ordens de Serviço ────────────────────────────────────────────────────────

async def _query_service_orders(session, aid: str, date_from=None, date_to=None, so_status=None, so_priority=None, category=None):
    conds = ["association_id = :aid"]
    p: dict = {"aid": aid}
    if date_from: conds.append("created_at::date >= :df"); p["df"] = date.fromisoformat(date_from)
    if date_to: conds.append("created_at::date <= :dt"); p["dt"] = date.fromisoformat(date_to)
    if so_status: conds.append("status = :st"); p["st"] = so_status
    if so_priority: conds.append("priority = :pr"); p["pr"] = so_priority
    if category: conds.append("category_name ILIKE :cat"); p["cat"] = f"%{category}%"
    w = " AND ".join(conds)
    rows = (await session.execute(text(f"""
        SELECT number, title, status, priority, area, category_name,
               service_impacted, org_responsible, requester_name, requester_phone,
               assigned_to_name, request_date::date, created_at::date,
               resolved_at::date, resolution_notes, cancellation_reason
        FROM service_orders WHERE {w} ORDER BY number
    """), p)).fetchall()
    cols = ["Nº","Título","Status","Prioridade","Área","Categoria","Serviço Afetado",
            "Org. Responsável","Solicitante","Telefone","Atribuído a","Data Solicitação",
            "Criado em","Resolvido em","Notas Resolução","Motivo Cancelamento"]
    return [dict(zip(cols, [_s(v) for v in r])) for r in rows]


@router.get("/service-orders/preview")
async def preview_service_orders(
    date_from: str | None = None, date_to: str | None = None,
    so_status: str | None = None, so_priority: str | None = None, category: str | None = None,
    current: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> list[dict]:
    return await _query_service_orders(session, str(current.association_id), date_from, date_to, so_status, so_priority, category)


@router.get("/service-orders")
async def export_service_orders(
    date_from: str | None = None, date_to: str | None = None,
    so_status: str | None = None, so_priority: str | None = None, category: str | None = None,
    current: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> Response:
    rows = await _query_service_orders(session, str(current.association_id), date_from, date_to, so_status, so_priority, category)
    wb, ws = _mk("Ordens de Serviço")
    cols = list(rows[0].keys()) if rows else []
    _headers(ws, cols)
    for r in rows: ws.append(list(r.values()))
    _widths(ws)
    return _xlsx(wb, "ordens.xlsx")


# ─── Mensalidades ─────────────────────────────────────────────────────────────

STATUS_PT = {"paid": "Pago", "pending": "Pendente", "overdue": "Em atraso", "agreement": "Acordo", "waived": "Isento"}

async def _query_mensalidades(session, aid: str, date_from=None, date_to=None, men_status=None, ref_month=None, include_delinquent: bool = False):
    from datetime import timedelta
    from app.services.mensalidade_service import MensalidadeService
    from uuid import UUID as _UUID

    conds = ["v.association_id = :aid"]
    p: dict = {"aid": aid}
    if date_from: conds.append("COALESCE(v.paid_at, v.due_date::timestamp)::date >= :df"); p["df"] = date.fromisoformat(date_from)
    if date_to: conds.append("COALESCE(v.paid_at, v.due_date::timestamp)::date <= :dt"); p["dt"] = date.fromisoformat(date_to)
    if men_status: conds.append("v.status = :st"); p["st"] = men_status
    if ref_month: conds.append("v.reference_month = :rm"); p["rm"] = ref_month
    w = " AND ".join(conds)
    rows = (await session.execute(text(f"""
        SELECT v.resident_name, v.reference_month, v.due_date,
               v.amount, v.status, v.paid_at::date,
               v.payment_method_name, v.origem,
               TRIM(CONCAT_WS(', ',
                   NULLIF(r.address_street,''), NULLIF(r.address_number,''),
                   NULLIF(r.address_complement,''), NULLIF(r.address_cep,'')
               )) AS endereco,
               r.phone_primary,
               v.resident_id::text
        FROM v_mensalidades_completas v
        JOIN residents r ON r.id = v.resident_id
        WHERE {w} AND r.status != 'suspended' ORDER BY v.reference_month, v.resident_name
    """), p)).fetchall()
    cols = ["Morador","Mês Referência","Vencimento","Valor (R$)","Status","Pago em","Forma Pagamento","Origem","Endereço","Telefone","_rid","Estado Pagamento"]
    _pending_statuses = {"Pendente", "Em atraso"}
    def row_dict(r):
        d = dict(zip(cols[:-1], [_s(v) for v in r]))
        d["Status"] = STATUS_PT.get(d["Status"], d["Status"])
        d["Origem"] = "Sistema" if d["Origem"] == "sistema" else "Migração"
        d["Estado Pagamento"] = "Mensalidade"
        return d
    result = [row_dict(r) for r in rows]

    if include_delinquent:
        svc = MensalidadeService(session)
        delinquent = await svc.list_delinquent(_UUID(aid))
        from collections import defaultdict
        grouped: dict = defaultdict(list)
        for d in delinquent:
            grouped[d["resident_id"]].append(d)

        # Remove individual pending/overdue rows for delinquent residents
        # (replaced below by a single consolidated row)
        delinquent_ids = set(str(k) for k in grouped)
        result = [
            r for r in result
            if not (r.get("_rid") in delinquent_ids and r["Status"] in _pending_statuses)
        ]

        for resident_id, items in grouped.items():
            total = sum(float(i["amount"]) for i in items)
            months = sorted(i["reference_month"] for i in items)
            tasks = " + ".join(
                f"{m} pendente" for m in months
            )
            first = items[0]
            result.append({
                "Morador": first["resident_name"],
                "Mês Referência": f"{months[0]} a {months[-1]}" if len(months) > 1 else months[0],
                "Vencimento": first["due_date"],
                "Valor (R$)": f"{total:.2f}",
                "Status": "Inadimplente",
                "Pago em": "—",
                "Forma Pagamento": tasks,
                "Origem": "—",
                "Endereço": f"{first.get('address_street','')} {first.get('address_number','')}".strip(),
                "Telefone": first.get("phone_primary") or "—",
                "_rid": str(resident_id),
                "Estado Pagamento": "Inadimplência",
            })

    # Strip internal key before returning
    for r in result:
        r.pop("_rid", None)

    return result


@router.get("/mensalidades/preview")
async def preview_mensalidades(
    date_from: str | None = None, date_to: str | None = None,
    men_status: str | None = None, ref_month: str | None = None,
    include_delinquent: bool = False,
    current: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> list[dict]:
    return await _query_mensalidades(session, str(current.association_id), date_from, date_to, men_status, ref_month, include_delinquent)


@router.get("/mensalidades")
async def export_mensalidades(
    date_from: str | None = None, date_to: str | None = None,
    men_status: str | None = None, ref_month: str | None = None,
    include_delinquent: bool = False,
    current: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> Response:
    rows = await _query_mensalidades(session, str(current.association_id), date_from, date_to, men_status, ref_month, include_delinquent)
    wb, ws = _mk("Mensalidades")
    cols = list(rows[0].keys()) if rows else []
    _headers(ws, cols)
    for r in rows: ws.append(list(r.values()))
    _widths(ws)
    return _xlsx(wb, "mensalidades.xlsx")


# ─── Entregas ─────────────────────────────────────────────────────────────────

async def _aids_for_report(session: AsyncSession, aid: str) -> list[str]:
    row = (await session.execute(
        text("SELECT chat_group FROM associations WHERE id = :aid"), {"aid": aid}
    )).fetchone()
    group = row[0] if row else None
    if group:
        rows = (await session.execute(
            text("SELECT id FROM associations WHERE chat_group = :g"), {"g": group}
        )).fetchall()
    else:
        rows = (await session.execute(
            text("SELECT id FROM associations WHERE id = :aid"), {"aid": aid}
        )).fetchall()
    return [str(r[0]) for r in rows]


async def _query_entregas(
    session: AsyncSession,
    aids: list[str],
    date_from: str | None,
    date_to: str | None,
    user_id: str | None,
    types: str | None,
) -> list[dict]:
    import json as _json

    df = date.fromisoformat(date_from) if date_from else None
    dt = date.fromisoformat(date_to) if date_to else None
    active = set(types.split(",")) if types else {"tarefas", "checklist", "comentarios", "os", "demandas"}

    users_map: dict = {}

    def _u(uid: str, uname: str) -> dict:
        if uid not in users_map:
            users_map[uid] = {
                "user_id": uid, "user_name": uname, "items": [],
                "by_type": {"tarefas": 0, "checklist": 0, "comentarios": 0, "os": 0, "demandas": 0},
            }
        return users_map[uid]

    def _add(uid: str, uname: str, itype: str, title: str, idate: str | None, ref: str | None = None) -> None:
        u = _u(uid, uname)
        u["items"].append({"type": itype, "title": title, "date": idate, "ref": ref})
        u["by_type"][itype] += 1

    base: dict = {"aids": aids}
    if df: base["df"] = df
    if dt: base["dt"] = dt
    if user_id: base["uid"] = user_id

    def _dc(alias: str) -> str:
        parts = []
        if df: parts.append(f"AND {alias}.updated_at::date >= :df")
        if dt: parts.append(f"AND {alias}.updated_at::date <= :dt")
        return " ".join(parts)

    def _dcc(alias: str) -> str:
        parts = []
        if df: parts.append(f"AND {alias}.created_at::date >= :df")
        if dt: parts.append(f"AND {alias}.created_at::date <= :dt")
        return " ".join(parts)

    uc = "AND u.id = :uid" if user_id else ""

    if "tarefas" in active:
        rows = (await session.execute(text(f"""
            SELECT u.id, u.full_name, t.title, t.updated_at::date, t.service_order_title
            FROM daily_tasks t
            JOIN users u ON u.id = COALESCE(t.assigned_to, t.created_by)
            WHERE t.association_id = ANY(:aids) AND t.status = 'done'
              {_dc('t')} {uc}
            ORDER BY t.updated_at DESC
        """), base)).fetchall()
        for r in rows:
            _add(str(r[0]), r[1], "tarefas", r[2], str(r[3]) if r[3] else None, r[4])

    if "checklist" in active:
        rows = (await session.execute(text(f"""
            SELECT u.id, u.full_name, t.title, t.checklist, t.updated_at::date
            FROM daily_tasks t
            JOIN users u ON u.id = COALESCE(t.assigned_to, t.created_by)
            WHERE t.association_id = ANY(:aids)
              AND jsonb_array_length(t.checklist) > 0
              {_dc('t')} {uc}
            ORDER BY t.updated_at DESC
        """), base)).fetchall()
        for r in rows:
            cl = r[3]
            if isinstance(cl, str):
                try: cl = _json.loads(cl)
                except: cl = []
            for item in (cl or []):
                if item.get("done"):
                    _add(str(r[0]), r[1], "checklist", item.get("text", "—"), str(r[4]) if r[4] else None, r[2])

    if "comentarios" in active:
        uc_c = "AND c.created_by = :uid" if user_id else ""
        rows = (await session.execute(text(f"""
            SELECT u.id, u.full_name, c.comment, c.created_at::date, t.title
            FROM daily_task_comments c
            JOIN users u ON u.id = c.created_by
            JOIN daily_tasks t ON t.id = c.task_id
            WHERE c.association_id = ANY(:aids)
              {_dcc('c')} {uc_c}
            ORDER BY c.created_at DESC
        """), base)).fetchall()
        for r in rows:
            _add(str(r[0]), r[1], "comentarios", (r[2] or "—")[:80], str(r[3]) if r[3] else None, r[4])

    if "os" in active:
        base_no_uid = {k: v for k, v in base.items() if k != "uid"}
        parts = []
        if df: parts.append("AND so.updated_at::date >= :df")
        if dt: parts.append("AND so.updated_at::date <= :dt")
        dc_os = " ".join(parts)
        rows = (await session.execute(text(f"""
            SELECT u.id, u.full_name, so.title, so.number, so.updated_at::date
            FROM service_orders so
            JOIN users u ON u.id = so.created_by
            WHERE so.association_id = ANY(:aids) AND so.status = 'resolved'
              {dc_os}
            ORDER BY so.updated_at DESC
        """), base_no_uid)).fetchall()
        for r in rows:
            _add(str(r[0]), r[1], "os", r[2], str(r[4]) if r[4] else None, f"OS #{r[3]}")

    if "demandas" in active:
        rows = (await session.execute(text(f"""
            SELECT u.id, u.full_name, d.title, d.updated_at::date, so.number
            FROM demands d
            JOIN users u ON u.id = COALESCE(d.assigned_to, d.created_by)
            LEFT JOIN service_orders so ON so.id = d.service_order_id
            WHERE d.association_id = ANY(:aids) AND d.status = 'concluido'
              {_dc('d')} {uc}
            ORDER BY d.updated_at DESC
        """), base)).fetchall()
        for r in rows:
            _add(str(r[0]), r[1], "demandas", r[2], str(r[3]) if r[3] else None, f"OS #{r[4]}" if r[4] else None)

    result = list(users_map.values())
    for u in result:
        u["total"] = len(u["items"])
    result.sort(key=lambda x: -x["total"])
    return result


@router.get("/entregas")
async def get_entregas(
    date_from: str | None = None,
    date_to: str | None = None,
    user_id: str | None = None,
    types: str | None = None,
    current: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> list[dict]:
    aids = await _aids_for_report(session, str(current.association_id))
    return await _query_entregas(session, aids, date_from, date_to, user_id, types)


def _entregas_pdf(data: list[dict], date_from: str | None, date_to: str | None) -> bytes:
    from fpdf import FPDF  # type: ignore

    period = f"{date_from or '—'} a {date_to or '—'}"
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Relatório de Entregas", ln=True, align="C")
    pdf.set_font("Helvetica", size=10)
    pdf.cell(0, 6, f"Período: {period}", ln=True, align="C")
    pdf.ln(6)

    TYPE_LABELS = {"tarefas": "Tarefa", "checklist": "Checklist", "comentarios": "Comentário", "os": "OS", "demandas": "Demanda"}

    for u in data:
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_fill_color(26, 63, 111)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(0, 8, f"  {u['user_name']}  ({u['total']} entregas)", ln=True, fill=True)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", size=9)
        for item in u["items"]:
            tipo = TYPE_LABELS.get(item["type"], item["type"])
            ref = f" [{item['ref']}]" if item.get("ref") else ""
            dt_str = f" · {item['date']}" if item.get("date") else ""
            line = f"  [{tipo}]{dt_str}  {item['title']}{ref}"
            pdf.multi_cell(0, 6, _ILLEGAL.sub("", line))
        pdf.ln(4)

    return bytes(pdf.output())


def _entregas_xlsx(data: list[dict]) -> bytes:
    wb = Workbook()

    ws_sum = wb.active
    ws_sum.title = "Resumo"
    _headers(ws_sum, ["Colaborador", "Total", "Tarefas", "Checklist", "Comentários", "OS", "Demandas"])
    for u in data:
        bt = u["by_type"]
        ws_sum.append([u["user_name"], u["total"], bt["tarefas"], bt["checklist"], bt["comentarios"], bt["os"], bt["demandas"]])
    _widths(ws_sum)

    ws_det = wb.create_sheet("Detalhes")
    _headers(ws_det, ["Colaborador", "Tipo", "Título", "Data", "Referência"])
    TYPE_LABELS = {"tarefas": "Tarefa", "checklist": "Checklist", "comentarios": "Comentário", "os": "OS", "demandas": "Demanda"}
    for u in data:
        for item in u["items"]:
            ws_det.append([
                u["user_name"],
                TYPE_LABELS.get(item["type"], item["type"]),
                _s(item["title"]),
                _s(item.get("date")),
                _s(item.get("ref")),
            ])
    _widths(ws_det)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.get("/entregas/export")
async def export_entregas(
    date_from: str | None = None,
    date_to: str | None = None,
    user_id: str | None = None,
    types: str | None = None,
    fmt: str = "excel",
    current: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> Response:
    aids = await _aids_for_report(session, str(current.association_id))
    data = await _query_entregas(session, aids, date_from, date_to, user_id, types)
    if fmt == "pdf":
        pdf_bytes = _entregas_pdf(data, date_from, date_to)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": 'attachment; filename="entregas.pdf"'},
        )
    xlsx_bytes = _entregas_xlsx(data)
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="entregas.xlsx"'},
    )


class EmailEntregasRequest(BaseModel):
    email: str
    date_from: str | None = None
    date_to: str | None = None
    user_id: str | None = None
    types: str | None = None


@router.post("/entregas/email")
async def email_entregas(
    body: EmailEntregasRequest,
    background_tasks: BackgroundTasks,
    current: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> dict:
    aids = await _aids_for_report(session, str(current.association_id))
    data = await _query_entregas(session, aids, body.date_from, body.date_to, body.user_id, body.types)
    pdf_bytes = _entregas_pdf(data, body.date_from, body.date_to)
    period = f"{body.date_from or '—'} a {body.date_to or '—'}"

    def _send() -> None:
        from app.services.email_service import send_email
        send_email(
            to=body.email,
            subject=f"Relatório de Entregas — {period}",
            html=f"<p>Segue em anexo o relatório de entregas do período <strong>{period}</strong>.</p><p style='color:#6b7280;font-size:13px'>APRXM — Sistema de Gestão Comunitária</p>",
            pdf_attachment=pdf_bytes,
            pdf_filename="entregas.pdf",
        )

    background_tasks.add_task(_send)
    return {"ok": True}
